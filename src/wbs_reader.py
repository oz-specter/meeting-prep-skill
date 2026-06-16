"""WBS (.xlsx) を読み取り WbsView を返す（最小版・座標台帳のみ）.

設計方針:
  - openpyxl(read_only=True, data_only=True) で開く（書き込み禁止・mtime不変）
  - ヘッダ検出は必須列を最も多く含む行を採用（A1/C1 reader と同方針）
  - タスクID/タスク名 のみ必須。担当/開始日/終了日/工程 は任意。
  - シート欠落・必須列欠落 → None（例外を投げない）
  - タスクID は str(...).strip() で正規化、重複は先勝ち
  - 前提タスク等の解釈は不要（座標台帳のみ）

複製元: C2 risk-response-skill/src/wbs_reader.py（D1 report-skill 経由）の逐語複製。
C2確定IF: WbsView(sheet, col_letter, rows[, _values]) / task_cellref / read_wbs / WBS_COLUMN_ALIASES / WBS_REQUIRED_COLUMNS
D3加算: WBS_COLUMN_ALIASES に "工程" を追加（任意列・read_wbs本体無改変＝C2挙動不変）。
        工程列は upcoming_task 生成／related_phases 突合（prep=G5/G6）で使用。--wbs は任意連携。
ロジック改変禁止（由来コメント・加算のみ）。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .model import CellRef

# ---------------------------------------------------------------------------
# 定数（確定 IF）
# ---------------------------------------------------------------------------

WBS_SHEET_ALIASES: list[str] = ["WBS", "WBSシート", "作業計画", "WBS一覧"]

# 正規列名 -> 別名リスト
WBS_COLUMN_ALIASES: dict[str, list[str]] = {
    "タスクID": ["タスクID", "ID", "WBS番号", "No", "No."],
    "タスク名": ["タスク名", "作業名", "名称", "タスク"],
    "担当":     ["担当", "担当者", "オーナー", "アサイン"],
    "開始日":   ["開始日", "開始", "着手日"],
    "終了日":   ["終了日", "終了", "完了日"],
    "工程":     ["工程", "フェーズ", "フェイズ", "工程名"],
}

WBS_REQUIRED_COLUMNS: list[str] = ["タスクID", "タスク名"]


# ---------------------------------------------------------------------------
# データクラス（確定 IF）
# ---------------------------------------------------------------------------

@dataclass
class WbsView:
    """WBS シートの座標台帳.

    sheet: 実シート名（例 "WBS"）
    col_letter: 正規列名 -> 列記号 (例 "C")
    rows: タスクID -> 実Excel行 (1始まり)
    """

    sheet: str
    col_letter: dict[str, str]
    rows: dict[str, int]

    def task_cellref(
        self, task_id: str, canonical_col: str, note: str = ""
    ) -> "CellRef | None":
        """既知 task_id かつ 列存在 → CellRef。未知 task_id または列なし → None。

        例外を投げない。value は当該セルの生値（rows から行番号を引いて算出）。
        sheet は実シート名。render() は "WBS!<列><行> = \"<値>\"" 形式になる。
        """
        excel_row = self.rows.get(task_id)
        if excel_row is None:
            return None
        letter = self.col_letter.get(canonical_col)
        if letter is None:
            return None
        # value はこのビューには保持しないため "" として返す（座標台帳の役割に限定）
        # → _values 辞書がある場合はそちらを参照
        value = self._values.get((task_id, canonical_col), "")
        return CellRef(
            sheet=self.sheet,
            column=letter,
            row=excel_row,
            value=value,
            note=note,
        )

    # セル値の台帳（task_id, canonical_col) -> value）
    # dataclass field として持つが確定IFのシグネチャには表れない（内部用）
    _values: dict[tuple[str, str], object] = field(
        default_factory=dict, repr=False
    )


# ---------------------------------------------------------------------------
# 内部ヘルパ
# ---------------------------------------------------------------------------

def _build_reverse_alias(aliases: dict[str, list[str]]) -> dict[str, str]:
    """正規名 -> [表記] から 表記 -> 正規名 の逆引きを作る."""
    rev: dict[str, str] = {}
    for canon, variants in aliases.items():
        for v in variants:
            rev[str(v).strip()] = canon
    return rev


def _find_header(
    grid: list[tuple],
    required: list[str],
    rev_col: dict[str, str],
) -> "tuple[int | None, dict[str, int] | None]":
    """ヘッダ行インデックス (0始まり) と 正規列名->列インデックス を返す.

    必須列を最も多く含む行を採用（A1/C1 reader と同方針）。見つからなければ (None, None)。
    """
    best: "tuple[int, int, dict[str, int]] | None" = None
    for r_idx, row in enumerate(grid):
        col_index: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            canon = rev_col.get(str(cell).strip())
            if canon and canon not in col_index:
                col_index[canon] = c_idx
        matched = sum(1 for c in required if c in col_index)
        if matched and (best is None or matched > best[0]):
            best = (matched, r_idx, col_index)
    if best is None:
        return None, None
    return best[1], best[2]


# ---------------------------------------------------------------------------
# 公開 API（確定 IF）
# ---------------------------------------------------------------------------

def read_wbs(path: "str | Path") -> "WbsView | None":
    """WBS .xlsx を openpyxl(read_only=True, data_only=True) で読み WbsView を返す.

    失敗条件（→ None を返す、例外を投げない）:
      - WBS_SHEET_ALIASES に一致するシートが無い
      - 必須列（タスクID・タスク名）が欠落

    Excel への書き込みは一切しない。
    """
    try:
        path = Path(path)
        rev_col = _build_reverse_alias(WBS_COLUMN_ALIASES)

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            # ---- シート解決 ----------------------------------------
            target_title: "str | None" = None
            for ws in wb.worksheets:
                if str(ws.title).strip() in WBS_SHEET_ALIASES:
                    target_title = ws.title
                    break

            if target_title is None:
                return None

            ws = wb[target_title]
            grid = list(ws.iter_rows(values_only=True))

            # ---- ヘッダ行検出 ----------------------------------------
            header_idx, col_index = _find_header(
                grid, WBS_REQUIRED_COLUMNS, rev_col
            )

            if header_idx is None or col_index is None:
                return None

            # ---- 必須列欠落チェック ----------------------------------------
            missing = [c for c in WBS_REQUIRED_COLUMNS if c not in col_index]
            if missing:
                return None

            # ---- 列記号の辞書を構築 ----------------------------------------
            col_letter: dict[str, str] = {}
            for canon, idx in col_index.items():
                col_letter[canon] = get_column_letter(idx + 1)

            # ---- 明細行をパースして rows / _values を構築 ----------------------------------------
            rows: dict[str, int] = {}          # タスクID -> 実Excel行
            values: dict[tuple[str, str], object] = {}  # (task_id, canonical_col) -> value

            header_row_0 = header_idx  # 0始まりインデックス
            for j, row in enumerate(grid[header_row_0 + 1:]):
                excel_row = header_row_0 + 1 + j + 1  # 実Excel行番号（1始まり）

                # タスクID 取得
                id_idx = col_index.get("タスクID")
                if id_idx is None or id_idx >= len(row):
                    continue
                raw_id = row[id_idx]
                if raw_id is None:
                    continue
                task_id = str(raw_id).strip()
                if not task_id:
                    continue

                # 重複は先勝ち
                if task_id in rows:
                    continue

                rows[task_id] = excel_row

                # 全検出列の値を記録
                for canon, c_idx in col_index.items():
                    if c_idx < len(row):
                        values[(task_id, canon)] = row[c_idx]

            view = WbsView(
                sheet=target_title,
                col_letter=col_letter,
                rows=rows,
                _values=values,
            )
            return view

        finally:
            wb.close()

    except Exception:
        # 予期しない例外もクラッシュさせず None 退避
        return None
