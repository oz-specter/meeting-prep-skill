"""議事録 (.xlsx) を読み取り MinutesRegister を返す（D1 report-skill/src/minutes_reader.py 逐語複製）.

由来: D1 report-skill/src/minutes_reader.py（C1 risklog-skill reader の台帳方式踏襲）を D3 へ逐語複製。
コードは改変しない（由来コメントのみ）。後段 prep.py（G5/G6）が確定IFに依存する。

確定IF:
  read_minutes(path) -> MinutesRegister | MissingColumnReport
  SheetView(sheet, header_row, col_letter).cellref(canonical_col, excel_row, value, note="") -> CellRef
  MeetingEntry / MinutesRegister（下記フィールド）
openpyxl(read_only=True, data_only=True)・書込禁止・now()/today()禁止・決定論。
"""
from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .config import (
    MINUTES_COLUMN_ALIASES,
    MINUTES_OPTIONAL_COLUMNS,
    MINUTES_REQUIRED_COLUMNS,
    MINUTES_SHEET_ALIASES,
)
from .model import CellRef, MissingColumnReport

# 出席・関連タスク・関連要件の分割パターン（C1分割規則: カンマ・読点・全角カンマ・空白）
_SPLIT_PATTERN = re.compile(r"[,、，\s]+")


@dataclass
class SheetView:
    """1シートの正規化済みビュー（ヘッダ位置・列記号の台帳）.

    C1 risklog-skill/src/reader.py の SheetView と同型。
    """

    sheet: str                       # 実シート名（例 "議事録"）
    header_row: int                  # ヘッダの実Excel行（1始まり）
    col_letter: dict[str, str]       # 正規列名 -> 列記号(例 "C")

    def cellref(
        self, canonical_col: str, excel_row: int, value: object, note: str = ""
    ) -> CellRef:
        """正規列名＋実Excel行から CellRef を起こす（列記号は col_letter 由来）."""
        return CellRef(
            self.sheet,
            self.col_letter[canonical_col],
            excel_row,
            value,
            note,
        )


@dataclass
class MeetingEntry:
    """正規化済み1議事録行."""

    meeting_date: datetime.date      # パース済（entries は常に有効日付＝必須）
    date_raw: object                 # 生値（CellRef value 用）
    date_str: str                    # "YYYY-MM-DD"（= meeting_date.isoformat()）
    meeting_type: str                # 会議体 生表記 strip（**正規化しない**・None/"-"→""）
    decision: str                    # 決定事項 strip（None/"-"→""。RK-xx はテキストのまま）
    attendees: list[str]             # 出席 分割（C1分割規則・None/"-"→[]）
    homework_content: str            # 宿題_内容 strip（None/"-"→""）
    homework_owner: str              # 宿題_担当 strip（None/"-"→""）
    deadline: "datetime.date | None" # 宿題_期限 パース済（空/パース不能は None）
    deadline_raw: object             # 生値
    deadline_str: str                # "YYYY-MM-DD" or ""（空/不能）
    deadline_unparseable: bool       # 非空でパース不能（行は残す）
    related_tasks: list[str] = field(default_factory=list)   # 関連タスク 分割
    related_reqs: list[str] = field(default_factory=list)    # 関連要件 分割
    excel_row: int = 0               # 実Excel行（CellRef 用）


@dataclass
class MinutesRegister:
    """read_minutes の成功時返り値."""

    entries: list[MeetingEntry]      # 日付パース可能行のみ・入力順
    view: SheetView
    unparseable_dates: list[str] = field(default_factory=list)      # "行<実Excel行>"
    unparseable_deadlines: list[str] = field(default_factory=list)  # "行<実Excel行>"

    def max_meeting_date(self) -> "datetime.date | None":
        """entries の最大日付を返す（entries 空なら None）."""
        return max((e.meeting_date for e in self.entries), default=None)


def _build_reverse_alias(aliases: dict[str, list[str]]) -> dict[str, str]:
    """正規名 -> [表記] の辞書から 表記 -> 正規名 の逆引きを作る."""
    rev: dict[str, str] = {}
    for canon, variants in aliases.items():
        for v in variants:
            rev[str(v).strip()] = canon
    return rev


def _find_header(
    grid: list[tuple],
    required: list[str],
    rev_col: dict[str, str],
) -> "tuple[int | None, dict[str, int] | None]":
    """ヘッダ行インデックス (0始まり) と正規列名->列インデックスを返す.

    必須列を最も多く含む行を採用する（C1 reader と同一ロジック）。
    見つからなければ (None, None)。
    """
    best: "tuple[int, int, dict[str, int]] | None" = None  # (matched, row_idx, col_index)
    for r_idx, row in enumerate(grid):
        col_index: dict[str, int] = {}
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            canon = rev_col.get(str(cell).strip())
            if canon and canon not in col_index:
                col_index[canon] = c_idx
        matched = sum(1 for c in required if c in col_index)
        if matched and (best is None or matched > best[0]):
            best = (matched, r_idx, col_index)
    if best is None:
        return None, None
    return best[1], best[2]


def _cell_value(row: tuple, col_index: dict[str, int], canon: str) -> object:
    """行タプルから正規列名の値を取り出す。列が無い場合は None。"""
    idx = col_index.get(canon)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _split_list(v: object) -> list[str]:
    """カンマ・読点・全角カンマ・空白区切りで分割し、空要素をスキップして返す.

    C1分割規則: `[,、，\\s]+`。`"-"` 及び空は []。
    """
    if v is None:
        return []
    s = str(v).strip()
    if not s or s == "-":
        return []
    parts = _SPLIT_PATTERN.split(s)
    return [p for p in parts if p]


def _clean(v: object) -> str:
    """文字列クリーニング: None→""・strip・"-"→""."""
    if v is None:
        return ""
    s = str(v).strip()
    if s == "-":
        return ""
    return s


def _parse_date(v: object) -> "tuple[datetime.date | None, bool]":
    """日付生値 -> (datetime.date | None, is_invalid).

    C1 検知日方式（spec §2.1）:
    - datetime.datetime -> .date()（datetime を date より先に判定）
    - datetime.date -> そのまま
    - 文字列は strip 後 "%Y-%m-%d"・"%Y/%m/%d" を試す
    - None・空・"-" -> (None, False)
    - 非空でどの形式も不能 -> (None, True)
    """
    if v is None:
        return None, False

    # datetime.datetime は datetime.date のサブクラスなので先に判定
    if isinstance(v, datetime.datetime):
        return v.date(), False

    if isinstance(v, datetime.date):
        return v, False

    if isinstance(v, str):
        s = v.strip()
        if not s or s == "-":
            return None, False
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(s, fmt).date(), False
            except ValueError:
                pass
        return None, True

    # その他の型（int など）は非空なのでパース不能として is_invalid=True
    return None, True


def read_minutes(path: "str | Path") -> "MinutesRegister | MissingColumnReport":
    """議事録 .xlsx を読み取り MinutesRegister を返す.

    失敗時（シート欠落・必須列欠落）は MissingColumnReport を返す。
    Excel本体への書き込みは一切行わない（openpyxl read_only=True, data_only=True）。

    Args:
        path: 対象 .xlsx のパス（str または Path）。

    Returns:
        MinutesRegister（成功）または MissingColumnReport（欠落）。
    """
    path = Path(path)
    rev_col = _build_reverse_alias(MINUTES_COLUMN_ALIASES)
    rev_sheet = _build_reverse_alias(MINUTES_SHEET_ALIASES)

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        # ---- シート解決 ----------------------------------------
        target_sheet_title: "str | None" = None
        for ws in wb.worksheets:
            canon = rev_sheet.get(str(ws.title).strip())
            if canon == "議事録":
                target_sheet_title = ws.title
                break

        if target_sheet_title is None:
            return MissingColumnReport(
                sheet="議事録",
                missing_columns=[],
                suggestion=(
                    "シート『議事録』が見つかりません。"
                    "シート名が MINUTES_SHEET_ALIASES に含まれているかご確認ください。"
                    "（Excel側の変更は不要です。MINUTES_SHEET_ALIASES に別名を追加できます）"
                ),
            )

        ws = wb[target_sheet_title]
        grid = list(ws.iter_rows(values_only=True))

        # ---- ヘッダ行検出 ----------------------------------------
        header_idx, col_index = _find_header(grid, MINUTES_REQUIRED_COLUMNS, rev_col)

        if header_idx is None or col_index is None:
            return MissingColumnReport(
                sheet=target_sheet_title,
                missing_columns=MINUTES_REQUIRED_COLUMNS[:],
                suggestion="ヘッダ行を検出できませんでした。列見出しの行をご確認ください。",
            )

        # ---- 必須列欠落チェック ----------------------------------------
        missing = [c for c in MINUTES_REQUIRED_COLUMNS if c not in col_index]
        if missing:
            return MissingColumnReport(
                sheet=target_sheet_title,
                missing_columns=missing,
                suggestion=(
                    f"『{target_sheet_title}』シートに必須列が不足しています: {missing}。"
                    "列の表記揺れなら config.MINUTES_COLUMN_ALIASES に別名を追加できます。"
                    "（Excel側の修正は不要、こちらで吸収します）"
                ),
            )

        # ---- col_letter 辞書を構築（必須＋任意列）----------------------------------------
        # MINUTES_OPTIONAL_COLUMNS のうち実際に存在する列も col_letter に含める
        all_known = set(MINUTES_REQUIRED_COLUMNS) | set(MINUTES_OPTIONAL_COLUMNS)
        col_letter: dict[str, str] = {}
        for canon, idx in col_index.items():
            if canon in all_known:
                col_letter[canon] = get_column_letter(idx + 1)

        view = SheetView(
            sheet=target_sheet_title,
            header_row=header_idx + 1,   # 0始まり -> 1始まり実Excel行
            col_letter=col_letter,
        )

        # ---- 明細行のパース ----------------------------------------
        unparseable_dates: list[str] = []
        unparseable_deadlines: list[str] = []
        entries: list[MeetingEntry] = []

        header_row_0 = header_idx  # 0始まりインデックス
        for j, row in enumerate(grid[header_row_0 + 1:]):
            excel_row = header_row_0 + 1 + j + 1   # 実Excel行番号（1始まり）

            date_raw = _cell_value(row, col_index, "日付")

            # --- 日付パース ---
            date_val, date_invalid = _parse_date(date_raw)

            # 日付が空（None,False かつ生値が空/"-"/None）-> 行スキップ・notice なし
            if date_val is None and not date_invalid:
                # 空扱い（None, 空文字, "-"）
                continue

            # 非空でパース不能 -> 行除外＋notice
            if date_val is None and date_invalid:
                unparseable_dates.append(f"行{excel_row}")
                continue

            # --- 任意列のパース ---
            meeting_type_raw = _cell_value(row, col_index, "会議体")
            meeting_type = _clean(meeting_type_raw)

            decision_raw = _cell_value(row, col_index, "決定事項")
            decision = _clean(decision_raw)

            attendees_raw = _cell_value(row, col_index, "出席")
            attendees = _split_list(attendees_raw)

            hw_content_raw = _cell_value(row, col_index, "宿題_内容")
            homework_content = _clean(hw_content_raw)

            hw_owner_raw = _cell_value(row, col_index, "宿題_担当")
            homework_owner = _clean(hw_owner_raw)

            # --- 宿題_期限のパース ---
            deadline_raw = _cell_value(row, col_index, "宿題_期限")
            deadline_val, deadline_invalid = _parse_date(deadline_raw)

            deadline_unparseable = False
            deadline_str = ""

            if deadline_val is not None:
                # パース成功
                deadline_str = deadline_val.isoformat()
            elif deadline_invalid:
                # 非空でパース不能 -> 行は残す・notice
                deadline_unparseable = True
                unparseable_deadlines.append(f"行{excel_row}")

            # --- 関連タスク・関連要件 ---
            rel_tasks_raw = _cell_value(row, col_index, "関連タスク")
            related_tasks = _split_list(rel_tasks_raw)

            rel_reqs_raw = _cell_value(row, col_index, "関連要件")
            related_reqs = _split_list(rel_reqs_raw)

            entries.append(MeetingEntry(
                meeting_date=date_val,  # type: ignore[arg-type]
                date_raw=date_raw,
                date_str=date_val.isoformat(),  # type: ignore[union-attr]
                meeting_type=meeting_type,
                decision=decision,
                attendees=attendees,
                homework_content=homework_content,
                homework_owner=homework_owner,
                deadline=deadline_val,
                deadline_raw=deadline_raw,
                deadline_str=deadline_str,
                deadline_unparseable=deadline_unparseable,
                related_tasks=related_tasks,
                related_reqs=related_reqs,
                excel_row=excel_row,
            ))

        return MinutesRegister(
            entries=entries,
            view=view,
            unparseable_dates=unparseable_dates,
            unparseable_deadlines=unparseable_deadlines,
        )

    finally:
        wb.close()
