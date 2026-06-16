"""tests/test_wbs_reader.py — WbsView / read_wbs の単体テスト（C2複製＋工程加算・AC-2）.

由来: C2 risk-response-skill/tests/test_wbs_reader.py（C3 escalation-skill 経由）逐語複製＋工程テスト。
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.model import CellRef
from src.wbs_reader import WbsView, read_wbs


class TestReadWbsSample:
    def test_returns_wbsview(self, wbs_sample_path: Path) -> None:
        assert isinstance(read_wbs(wbs_sample_path), WbsView)

    def test_rows_not_empty(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert len(view.rows) > 0

    def test_col_letter_contains_required(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert "タスクID" in view.col_letter
        assert "タスク名" in view.col_letter

    def test_sheet_name(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert view.sheet == "WBS"


class TestTaskCellrefKnown:
    def test_task_name_returns_cellref(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert isinstance(view.task_cellref("T01a", "タスク名"), CellRef)

    def test_task_name_render_contains_wbs_sheet(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        ref = view.task_cellref("T01a", "タスク名")
        assert isinstance(ref, CellRef)
        assert "WBS!" in ref.render()

    def test_owner_returns_cellref(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert isinstance(view.task_cellref("T01a", "担当"), CellRef)

    def test_start_date_returns_cellref(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert isinstance(view.task_cellref("T01a", "開始日"), CellRef)

    def test_end_date_returns_cellref(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert isinstance(view.task_cellref("T01a", "終了日"), CellRef)

    def test_cellref_has_correct_sheet(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        ref = view.task_cellref("T01a", "タスク名")
        assert isinstance(ref, CellRef)
        assert ref.sheet == "WBS"

    def test_cellref_column_from_col_letter(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        ref = view.task_cellref("T01a", "タスク名")
        assert isinstance(ref, CellRef)
        assert ref.column == view.col_letter["タスク名"]


class TestTaskCellrefNone:
    def test_unknown_task_id_returns_none(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert view.task_cellref("ZZZ-999", "タスク名") is None

    def test_unknown_column_returns_none(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        assert view.task_cellref("T01a", "存在しない列") is None

    def test_unknown_task_id_no_exception(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        try:
            view.task_cellref("NONEXISTENT-TASK-ID-12345", "タスク名")
        except Exception as e:
            pytest.fail(f"task_cellref が例外を投げた: {e!r}")

    def test_unknown_column_no_exception(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        try:
            view.task_cellref("T01a", "全く存在しない列XXXX")
        except Exception as e:
            pytest.fail(f"task_cellref が例外を投げた: {e!r}")


class TestMtimeUnchanged:
    def test_mtime_unchanged_after_read(self, wbs_sample_path: Path) -> None:
        before = os.stat(wbs_sample_path).st_mtime_ns
        read_wbs(wbs_sample_path)
        after = os.stat(wbs_sample_path).st_mtime_ns
        assert before == after


class TestMissingSheetOrColumn:
    def _make_xlsx_no_wbs_sheet(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["タスクID", "タスク名"])
        ws.append(["T01", "テストタスク"])
        out = tmp_path / "no_wbs_sheet.xlsx"
        wb.save(str(out))
        return out

    def _make_xlsx_no_taskid_column(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "WBS"
        ws.append(["作業名", "担当"])
        ws.append(["テストタスク", "担当者A"])
        out = tmp_path / "no_taskid_col.xlsx"
        wb.save(str(out))
        return out

    def _make_xlsx_no_taskname_column(self, tmp_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "WBS"
        ws.append(["タスクID", "担当"])
        ws.append(["T01", "担当者A"])
        out = tmp_path / "no_taskname_col.xlsx"
        wb.save(str(out))
        return out

    def test_no_wbs_sheet_returns_none(self, tmp_path: Path) -> None:
        assert read_wbs(self._make_xlsx_no_wbs_sheet(tmp_path)) is None

    def test_no_wbs_sheet_no_exception(self, tmp_path: Path) -> None:
        try:
            read_wbs(self._make_xlsx_no_wbs_sheet(tmp_path))
        except Exception as e:
            pytest.fail(f"read_wbs が例外を投げた: {e!r}")

    def test_no_taskid_column_returns_none(self, tmp_path: Path) -> None:
        assert read_wbs(self._make_xlsx_no_taskid_column(tmp_path)) is None

    def test_no_taskname_column_returns_none(self, tmp_path: Path) -> None:
        assert read_wbs(self._make_xlsx_no_taskname_column(tmp_path)) is None

    def test_no_taskid_column_no_exception(self, tmp_path: Path) -> None:
        try:
            read_wbs(self._make_xlsx_no_taskid_column(tmp_path))
        except Exception as e:
            pytest.fail(f"read_wbs が例外を投げた: {e!r}")

    def test_no_taskname_column_no_exception(self, tmp_path: Path) -> None:
        try:
            read_wbs(self._make_xlsx_no_taskname_column(tmp_path))
        except Exception as e:
            pytest.fail(f"read_wbs が例外を投げた: {e!r}")

    def test_no_wbs_sheet_sample_not_modified(self, tmp_path: Path) -> None:
        xlsx_path = self._make_xlsx_no_wbs_sheet(tmp_path)
        assert tmp_path in xlsx_path.parents
        assert read_wbs(xlsx_path) is None


class TestWbsViewContent:
    def test_t01a_row_is_integer(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        row = view.rows.get("T01a")
        assert isinstance(row, int) and row > 0

    def test_task_cellref_row_matches_rows_dict(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        expected_row = view.rows["T01a"]
        ref = view.task_cellref("T01a", "タスク名")
        assert isinstance(ref, CellRef)
        assert ref.row == expected_row

    def test_multiple_tasks_in_rows(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        for tid in ["T01a", "T02a", "T03a"]:
            assert tid in view.rows, f"タスクID {tid!r} が rows に含まれない"

    def test_note_parameter_passed_to_cellref(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        ref = view.task_cellref("T01a", "タスク名", note="テスト注記")
        assert isinstance(ref, CellRef)
        assert ref.note == "テスト注記"

    def test_render_format(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert isinstance(view, WbsView)
        ref = view.task_cellref("T01a", "タスク名")
        assert isinstance(ref, CellRef)
        import re
        assert re.match(r"WBS![A-Z]+\d+ = ", ref.render())


class TestPhaseColumn:
    """工程列の読み取りと任意列としての挙動を検証する（D3加算）."""

    def test_phase_column_detected(self, wbs_sample_path: Path) -> None:
        view = read_wbs(wbs_sample_path)
        assert view is not None
        assert "工程" in view.col_letter

    def test_phase_cellref_values(
        self, wbs_sample_path: Path, sample_wbs_phase_by_task: dict
    ) -> None:
        view = read_wbs(wbs_sample_path)
        assert view is not None
        for task_id, phase in sample_wbs_phase_by_task.items():
            cr = view.task_cellref(task_id, "工程")
            assert cr is not None, f"task_cellref({task_id!r}, '工程') が None"
            assert cr.value == phase, (
                f"task_id={task_id!r}: expected {phase!r}, got {cr.value!r}"
            )
            assert "WBS!" in cr.render()

    def test_phase_column_optional_when_absent(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "WBS"
        ws.append(["タスクID", "タスク名"])
        ws.append(["T01", "テストタスク1"])
        ws.append(["T02", "テストタスク2"])
        tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)
        try:
            wb.save(tmp_path_str)
            view = read_wbs(tmp_path_str)
            assert view is not None
            assert "工程" not in view.col_letter
        finally:
            os.remove(tmp_path_str)
