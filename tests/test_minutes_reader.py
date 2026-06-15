"""tests/test_minutes_reader.py — MinutesRegister の受け入れテスト（D1逐語複製・AC-1）.

spec §2 の全パース規則＋確定IFを機械検証する。
由来: D1 report-skill/tests/test_minutes_reader.py 逐語複製。
"""
from __future__ import annotations

import datetime
import os
from pathlib import Path

import openpyxl
import pytest

from src.minutes_reader import (
    MeetingEntry,
    MinutesRegister,
    SheetView,
    read_minutes,
)
from src.model import CellRef, MissingColumnReport


def _make_minutes_xlsx(
    tmp_path: Path,
    rows: list[tuple],
    sheet_name: str = "議事録",
    filename: str = "test_minutes.xlsx",
) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    out = tmp_path / filename
    wb.save(str(out))
    return out


class TestSampleMinutes:
    def test_returns_minutes_register(self, minutes_sample_path):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister), f"返り値が MinutesRegister でない: {type(reg)}"

    def test_entry_count(self, minutes_sample_path, sample_meeting_count):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == sample_meeting_count

    def test_action_count(self, minutes_sample_path, sample_action_count):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        action_count = sum(1 for e in reg.entries if e.homework_content)
        assert action_count == sample_action_count

    def test_max_meeting_date(self, minutes_sample_path):
        """reg.max_meeting_date() == 2026-07-10（D3 as_of アンカー A-1）."""
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        assert reg.max_meeting_date() == datetime.date(2026, 7, 10)

    def test_clean_notices(self, minutes_sample_path):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        assert reg.unparseable_dates == []
        assert reg.unparseable_deadlines == []

    def test_all_entries_have_date_type(self, minutes_sample_path):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        for e in reg.entries:
            assert isinstance(e.meeting_date, datetime.date)

    def test_cellref_contains_sheet_name(self, minutes_sample_path):
        reg = read_minutes(minutes_sample_path)
        assert isinstance(reg, MinutesRegister)
        for e in reg.entries:
            ref = reg.view.cellref("決定事項", e.excel_row, e.decision)
            assert "議事録!" in ref.render()

    def test_mtime_unchanged(self, minutes_sample_path):
        before = os.stat(str(minutes_sample_path)).st_mtime_ns
        read_minutes(minutes_sample_path)
        after = os.stat(str(minutes_sample_path)).st_mtime_ns
        assert before == after


class TestDateParsing:
    def test_empty_date_skips_row(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), (None, "週次定例", "テスト決定")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 0
        assert reg.unparseable_dates == []

    def test_invalid_date_excludes_row(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("foo", "週次定例", "テスト決定")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 0
        assert reg.unparseable_dates == ["行2"]

    def test_dash_date_skips_row(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("-", "週次定例", "テスト決定")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 0
        assert reg.unparseable_dates == []

    def test_date_formats(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "議事録"
        ws.append(("日付", "会議体", "決定事項"))
        ws.append(("2026-06-05", "会議A", "決定A"))
        ws.append(("2026/06/05", "会議B", "決定B"))
        ws.append((datetime.date(2026, 6, 5), "会議C", "決定C"))
        ws.append((datetime.datetime(2026, 6, 5), "会議D", "決定D"))
        out = tmp_path / "date_formats.xlsx"
        wb.save(str(out))
        reg = read_minutes(out)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 4
        for e in reg.entries:
            assert e.meeting_date == datetime.date(2026, 6, 5)


class TestDeadlineParsing:
    def _make_simple(self, tmp_path: Path, deadline_value) -> MinutesRegister:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "議事録"
        ws.append(("日付", "会議体", "決定事項", "宿題_内容", "宿題_担当", "宿題_期限"))
        ws.append(("2026-06-05", "週次定例", "決定A", "宿題A", "CN-01", deadline_value))
        out = tmp_path / "deadline_test.xlsx"
        wb.save(str(out))
        reg = read_minutes(out)
        assert isinstance(reg, MinutesRegister)
        return reg

    def test_empty_deadline_entry_remains(self, tmp_path):
        reg = self._make_simple(tmp_path, None)
        assert len(reg.entries) == 1
        assert reg.entries[0].deadline is None
        assert reg.unparseable_deadlines == []

    def test_bad_deadline_entry_remains(self, tmp_path):
        reg = self._make_simple(tmp_path, "bad")
        assert len(reg.entries) == 1
        e = reg.entries[0]
        assert e.deadline is None
        assert e.deadline_unparseable is True
        assert reg.unparseable_deadlines == ["行2"]

    def test_valid_deadline_slash(self, tmp_path):
        reg = self._make_simple(tmp_path, "2026/07/01")
        assert len(reg.entries) == 1
        assert reg.entries[0].deadline == datetime.date(2026, 7, 1)
        assert reg.entries[0].deadline_str == "2026-07-01"
        assert reg.unparseable_deadlines == []


class TestSplitList:
    def test_attendees_split(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項", "出席"), ("2026-06-05", "定例", "決定", "A, B、C，D E")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].attendees == ["A", "B", "C", "D", "E"]

    def test_attendees_dash(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項", "出席"), ("2026-06-05", "定例", "決定", "-")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].attendees == []

    def test_related_tasks_split(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項", "関連タスク"), ("2026-06-05", "定例", "決定", "T01,T02、T03")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].related_tasks == ["T01", "T02", "T03"]

    def test_related_reqs_split(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項", "関連要件"), ("2026-06-05", "定例", "決定", "R-01 R-02")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].related_reqs == ["R-01", "R-02"]


class TestMeetingTypeNormalization:
    def test_meeting_type_not_normalized(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [
                ("日付", "会議体", "決定事項"),
                ("2026-06-05", "週次定例", "決定A"),
                ("2026-06-12", "定例", "決定B"),
            ],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 2
        types = [e.meeting_type for e in reg.entries]
        assert "週次定例" in types
        assert "定例" in types


class TestDecisionRKPreservation:
    def test_rk_ref_preserved_in_decision(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("2026-06-05", "定例", "RK-05 対応を承認")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert "RK-05" in reg.entries[0].decision


class TestCleaning:
    def test_decision_none_becomes_empty(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("2026-06-05", "定例", None)],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].decision == ""

    def test_decision_dash_becomes_empty(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("2026-06-05", "定例", "-")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].decision == ""

    def test_homework_owner_strip(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項", "宿題_担当"), ("2026-06-05", "定例", "決定", None)],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert reg.entries[0].homework_owner == ""


class TestMissingColumnReport:
    def test_missing_required_column(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体"), ("2026-06-05", "定例")],
        )
        result = read_minutes(xlsx)
        assert isinstance(result, MissingColumnReport)
        assert "決定事項" in result.missing_columns

    def test_wrong_sheet_name(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("2026-06-05", "定例", "決定")],
            sheet_name="謎シート",
        )
        result = read_minutes(xlsx)
        assert isinstance(result, MissingColumnReport)
        assert result.sheet == "議事録"


class TestExcelRow:
    def test_excel_row_matches_actual_row(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [
                ("日付", "会議体", "決定事項"),
                ("2026-06-05", "定例A", "決定A"),
                ("2026-06-12", "定例B", "決定B"),
            ],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        assert len(reg.entries) == 2
        assert reg.entries[0].excel_row == 2
        assert reg.entries[1].excel_row == 3

    def test_cellref_row_via_view(self, tmp_path):
        xlsx = _make_minutes_xlsx(
            tmp_path,
            [("日付", "会議体", "決定事項"), ("2026-06-05", "定例", "決定A")],
        )
        reg = read_minutes(xlsx)
        assert isinstance(reg, MinutesRegister)
        e = reg.entries[0]
        ref = reg.view.cellref("決定事項", e.excel_row, e.decision)
        assert isinstance(ref, CellRef)
        assert ref.row == e.excel_row
